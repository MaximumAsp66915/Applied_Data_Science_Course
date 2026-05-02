import asyncio
import re
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

import aiohttp
from bs4 import BeautifulSoup
import json

BASE_URL = "https://bama.ir/car/samand?year=1385-2006"

HEADERS = {
    "User-Agent": "Mozilla/5.0"
}

def save_to_excel(results, filename="cars.xlsx"):
    wb = Workbook()
    ws = wb.active
    ws.title = "Cars"

    # Headers
    headers = [
        "URL", "Price", "Mileage", "Color",
        "Production Year", "Transmission", "Description"
    ]

    ws.append(headers)

    # Style headers
    for col in ws[1]:
        col.font = Font(bold=True)
        col.alignment = Alignment(horizontal="center")

    # Add data
    for row in results:
        url, price, mileage, color, year, transmission, description = row

        ws.append([
            url,
            price,
            mileage,
            color,
            year,
            transmission,
            description
        ])

    # Auto column width
    for column in ws.columns:
        max_length = 0
        col_letter = column[0].column_letter

        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass

        ws.column_dimensions[col_letter].width = min(max_length + 2, 50)

    ws.freeze_panes = "A2"
    for row in ws.iter_rows(min_row=2):
        row[6].alignment = Alignment(wrap_text=True)

    wb.save(filename)
    print(f"\n✅ Saved to {filename}")


async def get_car_details(session, url, semaphore):
    async with semaphore:
        try:
            async with session.get(url, timeout=10) as response:
                if response.status != 200:
                    return url, None, None, None, None, None, None

                html = await response.text()
                soup = BeautifulSoup(html, 'html.parser')

                scripts = soup.find_all('script', type='application/ld+json')

                # Default values
                price = None
                mileage = None
                color = None
                production_year = None
                transmission = None
                description = None

                for script in scripts:
                    try:
                        data = json.loads(script.string)

                        if not isinstance(data, dict):
                            continue

                        # Top-level
                        description = data.get("description") or description
                        color = data.get("color") or color

                        offers = data.get("offers", {})
                        item = offers.get("itemOffered", {})

                        # Price
                        price = offers.get("price") or price

                        # Mileage (sometimes nested deeper)

                        mileage_data = item.get("mileageFromOdometer", {})
                        if isinstance(mileage_data, dict):
                            mileage = mileage_data.get("value") or mileage

                        # Fallback from HTML
                        if mileage is None:
                            mileage_tag = soup.find(string=lambda t: t and "کیلومتر" in t)
                            if mileage_tag:
                                text = mileage_tag.strip()
                                match = re.search(r'[\d,]+', text)
                                mileage = match.group().replace(',', '') if match else text

                        if isinstance(mileage_data, dict):
                            mileage = mileage_data.get("value") or mileage

                        # Color fallback
                        color = item.get("color") or color

                        # Production year
                        production_year = item.get("productionDate") or production_year

                        # Transmission
                        transmission = item.get("vehicleTransmission") or transmission

                    except Exception:
                        continue

                return url, price, mileage, color, production_year, transmission, description

        except Exception:
            return url, None, None, None, None, None, None


async def main():
    semaphore = asyncio.Semaphore(10)

    async with aiohttp.ClientSession(headers=HEADERS) as session:

        async with session.get(BASE_URL) as response:
            listing_html = await response.text()
            soup = BeautifulSoup(listing_html, 'html.parser')

            links = []
            for a in soup.find_all('a', href=True):
                href = a['href']
                if "/car/detail-" in href:
                    full_link = f"https://bama.ir{href}" if href.startswith('/') else href
                    links.append(full_link)

        if not links:
            print("No links found.")
            return

        print(f"Found {len(links)} links. Starting extraction...\n")

        tasks = [get_car_details(session, link, semaphore) for link in links]
        results = await asyncio.gather(*tasks)

        print("------ RESULTS ------\n")

        for url, price, mileage, color, year, transmission, description in results:
            print(f"""
URL: {url}
Price: {price}
Mileage: {mileage}
Color: {color}
Production Year: {year}
Transmission: {transmission}
Description: {description[:100] + '...' if description else None}
---------------------------
""")

        from datetime import datetime
        filename = f"cars_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        save_to_excel(results, filename)

if __name__ == "__main__":
    asyncio.run(main())